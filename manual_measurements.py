#!/usr/bin/env python3.12
"""
Compute bone length measurements from manually labelled keypoints (labels.json).
Produces the same Excel format as predict_xray_images.py for direct comparison.
Only bones 1-3 per finger, matching model_11 predictions.

Output: manual_bone_measurements.xlsx
"""

import json
import numpy as np
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import (PatternFill, Font, Alignment, Border, Side)
from openpyxl.utils import get_column_letter

# =============================================================================
# CONFIG
# =============================================================================

CONFIG = {
    'annotations_file': '../labels.json',
    'px_to_mm':         0.265,
    'output_file':      'manual_bone_measurements.xlsx',
}

BONES = [
    ('L21', 'L21-', 'L21_'),
    ('L22', 'L22-', 'L22_'),
    ('L23', 'L23-', 'L23_'),
    ('L41', 'L41-', 'L41_'),
    ('L42', 'L42-', 'L42_'),
    ('L43', 'L43-', 'L43_'),
    ('R21', 'R21-', 'R21_'),
    ('R22', 'R22-', 'R22_'),
    ('R23', 'R23-', 'R23_'),
    ('R41', 'R41-', 'R41_'),
    ('R42', 'R42-', 'R42_'),
    ('R43', 'R43-', 'R43_'),
]

FINGER_BONE_GROUPS = {
    'L2': ['L21', 'L22', 'L23'],
    'L4': ['L41', 'L42', 'L43'],
    'R2': ['R21', 'R22', 'R23'],
    'R4': ['R41', 'R42', 'R43'],
}

# =============================================================================
# EXCEL STYLES
# =============================================================================

def make_styles():
    hdr_fill   = PatternFill('solid', fgColor='2F4F8F')
    hdr_font   = Font(bold=True, color='FFFFFF', size=10)
    num_font   = Font(size=10)
    total_font = Font(bold=True, size=10)
    center     = Alignment(horizontal='center', vertical='center')
    thin       = Side(style='thin', color='AAAAAA')
    border     = Border(left=thin, right=thin, top=thin, bottom=thin)
    thick_side = Side(style='medium', color='555555')
    total_border = Border(left=thick_side, right=thick_side,
                          top=thick_side,  bottom=thick_side)

    grp_fills = {
        'L2': PatternFill('solid', fgColor='D6E4F7'),
        'L4': PatternFill('solid', fgColor='D6F7E4'),
        'R2': PatternFill('solid', fgColor='FFF3CD'),
        'R4': PatternFill('solid', fgColor='FAD7D7'),
    }
    total_fills = {
        'L2': PatternFill('solid', fgColor='A8C8EE'),
        'L4': PatternFill('solid', fgColor='A8EEC8'),
        'R2': PatternFill('solid', fgColor='FFE08A'),
        'R4': PatternFill('solid', fgColor='F4A8A8'),
    }
    return (hdr_fill, hdr_font, num_font, total_font, center,
            border, total_border, grp_fills, total_fills)

# =============================================================================
# BUILD EXCEL
# =============================================================================

def build_excel(rows, output_path):
    (hdr_fill, hdr_font, num_font, total_font, center,
     border, total_border, grp_fills, total_fills) = make_styles()

    wb = Workbook()
    ws = wb.active
    ws.title = "Bone Measurements"

    # Group rows by image
    from collections import OrderedDict
    rows_by_image = OrderedDict()
    for r in rows:
        rows_by_image.setdefault(r['image'], []).append(r)

    # ── Sheet 1: Bone Measurements ────────────────────────────────────────────
    headers = ['Image', 'Bone', 'Top X', 'Top Y', 'Bottom X', 'Bottom Y',
               'Length (px)', 'Length (mm)', 'Detected']
    for col, h in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = hdr_fill and hdr_font
        cell.fill = hdr_fill
        cell.alignment = center
        cell.border = border

    r_idx = 2
    for img_name, img_rows in rows_by_image.items():
        by_bone = {r['bone_id']: r for r in img_rows}
        for finger in ['L2', 'L4', 'R2', 'R4']:
            fill = grp_fills[finger]
            for bone_id in FINGER_BONE_GROUPS[finger]:
                r = by_bone.get(bone_id)
                if r is None:
                    continue
                vals = [
                    r['image'], r['bone_id'],
                    round(r['top_x'], 1), round(r['top_y'], 1),
                    round(r['bot_x'], 1), round(r['bot_y'], 1),
                    round(r['length_px'], 2), round(r['length_mm'], 2),
                    'Yes' if r['detected'] else 'No',
                ]
                for col, val in enumerate(vals, start=1):
                    cell = ws.cell(row=r_idx, column=col, value=val)
                    cell.fill = fill
                    cell.border = border
                    cell.alignment = center
                    cell.font = num_font
                r_idx += 1

            # Finger subtotal row
            det = [by_bone[b] for b in FINGER_BONE_GROUPS[finger]
                   if b in by_bone and by_bone[b]['detected']]
            total_px = sum(r['length_px'] for r in det)
            total_mm = sum(r['length_mm'] for r in det)
            t_fill = total_fills[finger]
            t_vals = [img_name, f'{finger} TOTAL', '', '', '', '',
                      round(total_px, 2), round(total_mm, 2), f'{len(det)}/3']
            for col, val in enumerate(t_vals, start=1):
                cell = ws.cell(row=r_idx, column=col, value=val)
                cell.fill = t_fill
                cell.border = total_border
                cell.alignment = center
                cell.font = total_font
            r_idx += 1

    # ── Sheet 2: Summary per bone ─────────────────────────────────────────────
    ws2 = wb.create_sheet("Summary per bone")
    bone_lengths = {b[0]: [] for b in BONES}
    for r in rows:
        if r['detected']:
            bone_lengths[r['bone_id']].append(r['length_mm'])

    hdr2 = ['Bone', 'N images', 'Mean (mm)', 'Std (mm)', 'Min (mm)', 'Max (mm)']
    for col, h in enumerate(hdr2, start=1):
        cell = ws2.cell(row=1, column=col, value=h)
        cell.font = hdr_font; cell.fill = hdr_fill
        cell.alignment = center; cell.border = border

    r2 = 2
    for finger in ['L2', 'L4', 'R2', 'R4']:
        for bone_id in FINGER_BONE_GROUPS[finger]:
            vals = bone_lengths[bone_id]
            fill = grp_fills[finger]
            row_vals = [
                bone_id, len(vals),
                round(float(np.mean(vals)), 2) if vals else None,
                round(float(np.std(vals)),  2) if vals else None,
                round(float(np.min(vals)),  2) if vals else None,
                round(float(np.max(vals)),  2) if vals else None,
            ]
            for col, val in enumerate(row_vals, start=1):
                cell = ws2.cell(row=r2, column=col, value=val)
                cell.fill = fill; cell.border = border
                cell.alignment = center; cell.font = num_font
            r2 += 1

        # Finger total summary row
        all_vals = [bone_lengths[b] for b in FINGER_BONE_GROUPS[finger]]
        n_images = len(all_vals[0]) if all_vals[0] else 0
        finger_totals = [
            sum(all_vals[b][i] for b in range(3) if i < len(all_vals[b]))
            for i in range(n_images)
        ]
        t_fill = total_fills[finger]
        t_vals = [
            f'{finger} TOTAL', len(finger_totals),
            round(float(np.mean(finger_totals)), 2) if finger_totals else None,
            round(float(np.std(finger_totals)),  2) if finger_totals else None,
            round(float(np.min(finger_totals)),  2) if finger_totals else None,
            round(float(np.max(finger_totals)),  2) if finger_totals else None,
        ]
        for col, val in enumerate(t_vals, start=1):
            cell = ws2.cell(row=r2, column=col, value=val)
            cell.fill = t_fill; cell.border = total_border
            cell.alignment = center; cell.font = total_font
        r2 += 1

    # ── Sheet 3: Finger Totals per image ─────────────────────────────────────
    ws3 = wb.create_sheet("Finger Totals per image")
    hdr3 = ['Image',
            'L2 total (px)', 'L2 total (mm)',
            'L4 total (px)', 'L4 total (mm)',
            'R2 total (px)', 'R2 total (mm)',
            'R4 total (px)', 'R4 total (mm)',
            'All 4 fingers (mm)']
    for col, h in enumerate(hdr3, start=1):
        cell = ws3.cell(row=1, column=col, value=h)
        cell.font = hdr_font; cell.fill = hdr_fill
        cell.alignment = center; cell.border = border

    for r3_idx, (img_name, img_rows) in enumerate(rows_by_image.items(), start=2):
        by_bone = {r['bone_id']: r for r in img_rows}
        row_vals = [img_name]
        grand_mm = 0.0
        for finger in ['L2', 'L4', 'R2', 'R4']:
            det = [by_bone[b] for b in FINGER_BONE_GROUPS[finger]
                   if b in by_bone and by_bone[b]['detected']]
            total_px = sum(r['length_px'] for r in det)
            total_mm = sum(r['length_mm'] for r in det)
            grand_mm += total_mm
            row_vals += [round(total_px, 2) if det else None,
                         round(total_mm, 2) if det else None]
        row_vals.append(round(grand_mm, 2))
        for col, val in enumerate(row_vals, start=1):
            cell = ws3.cell(row=r3_idx, column=col, value=val)
            cell.border = border; cell.alignment = center; cell.font = num_font

    # ── Sheet 4: 2D4D Ratios ──────────────────────────────────────────────────
    ws4 = wb.create_sheet("2D4D Ratios")
    hdr4 = ['Image', 'L2 total (px)', 'L4 total (px)', 'Left 2D:4D',
             'R2 total (px)', 'R4 total (px)', 'Right 2D:4D']
    for col, h in enumerate(hdr4, start=1):
        cell = ws4.cell(row=1, column=col, value=h)
        cell.font = hdr_font; cell.fill = hdr_fill
        cell.alignment = center; cell.border = border

    ratio_data = []
    for r4_idx, (img_name, img_rows) in enumerate(rows_by_image.items(), start=2):
        by_bone = {r['bone_id']: r for r in img_rows}
        l2_px = sum(by_bone[b]['length_px'] for b in FINGER_BONE_GROUPS['L2']
                    if b in by_bone and by_bone[b]['detected'])
        l4_px = sum(by_bone[b]['length_px'] for b in FINGER_BONE_GROUPS['L4']
                    if b in by_bone and by_bone[b]['detected'])
        r2_px = sum(by_bone[b]['length_px'] for b in FINGER_BONE_GROUPS['R2']
                    if b in by_bone and by_bone[b]['detected'])
        r4_px = sum(by_bone[b]['length_px'] for b in FINGER_BONE_GROUPS['R4']
                    if b in by_bone and by_bone[b]['detected'])
        left_ratio  = round(l2_px / l4_px, 4) if l4_px > 0 else None
        right_ratio = round(r2_px / r4_px, 4) if r4_px > 0 else None
        ratio_data.append((left_ratio, right_ratio))
        row_vals = [img_name,
                    round(l2_px, 2), round(l4_px, 2), left_ratio,
                    round(r2_px, 2), round(r4_px, 2), right_ratio]
        for col, val in enumerate(row_vals, start=1):
            cell = ws4.cell(row=r4_idx, column=col, value=val)
            cell.border = border; cell.alignment = center; cell.font = num_font

    n_rows = len(rows_by_image) + 2
    for label, func in [('Mean', np.mean), ('Std', np.std),
                         ('Min', np.min),  ('Max', np.max)]:
        left_vals  = [d[0] for d in ratio_data if d[0] is not None]
        right_vals = [d[1] for d in ratio_data if d[1] is not None]
        stat_row = [label, '', '',
                    round(float(func(left_vals)),  4) if left_vals  else None,
                    '', '',
                    round(float(func(right_vals)), 4) if right_vals else None]
        for col, val in enumerate(stat_row, start=1):
            cell = ws4.cell(row=n_rows, column=col, value=val)
            cell.fill = hdr_fill; cell.border = total_border
            cell.alignment = center; cell.font = total_font
        n_rows += 1

    # ── Column widths ─────────────────────────────────────────────────────────
    for ws_ in [ws, ws2, ws3, ws4]:
        for col in ws_.columns:
            max_len = max((len(str(c.value)) if c.value is not None else 0)
                          for c in col)
            ws_.column_dimensions[
                get_column_letter(col[0].column)].width = max(max_len + 3, 12)

    wb.save(output_path)

# =============================================================================
# MAIN
# =============================================================================

def main():
    with open(CONFIG['annotations_file']) as f:
        all_data = json.load(f)

    images_with_points = [d for d in all_data if d.get('points')]
    print(f"Found {len(images_with_points)} images with manual annotations.\n")

    all_rows = []
    detected_total = 0

    for img_data in images_with_points:
        points = {p['label']: (p['x'], p['y']) for p in img_data.get('points', [])}

        for bone_id, top_lbl, bot_lbl in BONES:
            top = points.get(top_lbl)
            bot = points.get(bot_lbl)
            detected = top is not None and bot is not None

            if detected:
                length_px = float(np.hypot(bot[0] - top[0], bot[1] - top[1]))
                length_mm = length_px * CONFIG['px_to_mm']
                detected_total += 1
            else:
                length_px = length_mm = 0.0

            all_rows.append({
                'image':     img_data['name'],
                'bone_id':   bone_id,
                'top_x':     top[0] if top else 0.0,
                'top_y':     top[1] if top else 0.0,
                'bot_x':     bot[0] if bot else 0.0,
                'bot_y':     bot[1] if bot else 0.0,
                'length_px': length_px,
                'length_mm': length_mm,
                'detected':  detected,
            })

    build_excel(all_rows, CONFIG['output_file'])

    total_possible = len(images_with_points) * len(BONES)
    print(f"Done. {detected_total}/{total_possible} bone measurements found.")
    print(f"Saved: {CONFIG['output_file']}")


if __name__ == '__main__':
    try:
        main()
    except Exception:
        import traceback; traceback.print_exc()
