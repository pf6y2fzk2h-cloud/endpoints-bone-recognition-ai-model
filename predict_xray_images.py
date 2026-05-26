#!/usr/bin/env python3.12
"""
Model-11 bone-length measurement on all images in xray_images/.
================================================================
Runs the full cascade (seg → crop → keypoint) on every PNG in
../xray_images/, calculates the pixel and mm distance between the
top (-) and bottom (_) endpoint of each bone, and saves results
to an Excel file.

Bones predicted (bones 1-3 per finger, 12 bones total):
  L21, L22, L23  |  L41, L42, L43
  R21, R22, R23  |  R41, R42, R43

Run:
    python3.12 predict_xray_images.py
"""

import sys; sys.stdout.reconfigure(line_buffering=True)
import torch
import torch.nn as nn
import torch.nn.functional as F
import torchvision.models as tv_models
from PIL import Image
import numpy as np
import cv2
import os
from pathlib import Path
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# =============================================================================
# CONFIG
# =============================================================================

CONFIG = {
    'seg_model_path': 'trained_model/best_seg_model.pth',
    'kp_model_path':  'trained_model/best_kp_model.pth',
    'input_folder':   '../xray_images',
    'output_file':    'bone_measurements_xray_images.xlsx',
    'image_size':     256,
    'crop_size':      256,
    'seg_threshold':  0.2,
    'kp_threshold':   0.0,
    'pad_fraction':   0.10,
    'px_to_mm':       0.265,   # 96 DPI — 1 px = 0.265 mm
}

FINGER_LABELS   = ['L2', 'L4', 'R2', 'R4']
FINGER_KP_NAMES = {
    'L2': ['L21-','L21_','L22-','L22_','L23-','L23_'],
    'L4': ['L41-','L41_','L42-','L42_','L43-','L43_'],
    'R2': ['R21-','R21_','R22-','R22_','R23-','R23_'],
    'R4': ['R41-','R41_','R42-','R42_','R43-','R43_'],
}
ALL_KP_LABELS = (FINGER_KP_NAMES['L2'] + FINGER_KP_NAMES['L4'] +
                 FINGER_KP_NAMES['R2'] + FINGER_KP_NAMES['R4'])

# Bones: each consecutive pair (-, _) within a finger
BONES = [
    ('L21', 'L21-', 'L21_'), ('L22', 'L22-', 'L22_'), ('L23', 'L23-', 'L23_'),
    ('L41', 'L41-', 'L41_'), ('L42', 'L42-', 'L42_'), ('L43', 'L43-', 'L43_'),
    ('R21', 'R21-', 'R21_'), ('R22', 'R22-', 'R22_'), ('R23', 'R23-', 'R23_'),
    ('R41', 'R41-', 'R41_'), ('R42', 'R42-', 'R42_'), ('R43', 'R43-', 'R43_'),
]

# =============================================================================
# MODELS
# =============================================================================

class ConvBlock(nn.Module):
    def __init__(self, in_ch, out_ch):
        super().__init__()
        self.block = nn.Sequential(
            nn.Conv2d(in_ch, out_ch, 3, padding=1, bias=False),
            nn.BatchNorm2d(out_ch), nn.ReLU(inplace=True),
            nn.Conv2d(out_ch, out_ch, 3, padding=1, bias=False),
            nn.BatchNorm2d(out_ch), nn.ReLU(inplace=True),
        )
    def forward(self, x): return self.block(x)

class Up(nn.Module):
    def __init__(self, in_ch, skip_ch, out_ch):
        super().__init__()
        self.up   = nn.ConvTranspose2d(in_ch, in_ch // 2, 2, stride=2)
        self.conv = ConvBlock(in_ch // 2 + skip_ch, out_ch)
    def forward(self, x, skip):
        x = self.up(x)
        if x.shape[-2:] != skip.shape[-2:]:
            x = F.interpolate(x, size=skip.shape[-2:], mode='bilinear', align_corners=False)
        return self.conv(torch.cat([x, skip], dim=1))

class FingerSegUNet(nn.Module):
    def __init__(self, num_classes=4):
        super().__init__()
        bb = tv_models.resnet18(weights=None)
        self.enc0 = nn.Sequential(
            nn.Conv2d(1, 64, kernel_size=7, stride=2, padding=3, bias=False),
            bb.bn1, nn.ReLU(inplace=True))
        self.pool = bb.maxpool
        self.enc1 = bb.layer1; self.enc2 = bb.layer2
        self.enc3 = bb.layer3; self.enc4 = bb.layer4
        self.up4 = Up(512, 256, 256); self.up3 = Up(256, 128, 128)
        self.up2 = Up(128,  64,  64); self.up1 = Up(64,   64,  32)
        self.up0 = nn.Sequential(
            nn.ConvTranspose2d(32, 32, 2, stride=2), ConvBlock(32, 32))
        self.head = nn.Conv2d(32, num_classes, 1)
    def forward(self, x):
        e0 = self.enc0(x); e1 = self.enc1(self.pool(e0))
        e2 = self.enc2(e1); e3 = self.enc3(e2); b = self.enc4(e3)
        d = self.up4(b, e3); d = self.up3(d, e2); d = self.up2(d, e1)
        d = self.up1(d, e0); d = self.up0(d)
        return self.head(d)

class FingerKpModel(nn.Module):
    def __init__(self, num_kp=6):
        super().__init__()
        self.enc1   = ConvBlock(1,   32); self.pool1 = nn.MaxPool2d(2)
        self.enc2   = ConvBlock(32,  64); self.pool2 = nn.MaxPool2d(2)
        self.enc3   = ConvBlock(64, 128); self.pool3 = nn.MaxPool2d(2)
        self.bottle = ConvBlock(128, 256)
        self.up3    = nn.ConvTranspose2d(256, 128, 2, stride=2); self.dec3 = ConvBlock(256, 128)
        self.up2    = nn.ConvTranspose2d(128,  64, 2, stride=2); self.dec2 = ConvBlock(128,  64)
        self.up1    = nn.ConvTranspose2d( 64,  32, 2, stride=2); self.dec1 = ConvBlock( 64,  32)
        self.head   = nn.Conv2d(32, num_kp, 1)
    @staticmethod
    def _up(layer, x, ref):
        x = layer(x)
        if x.shape[-2:] != ref.shape[-2:]:
            x = F.interpolate(x, size=ref.shape[-2:], mode='bilinear', align_corners=False)
        return x
    def forward(self, x):
        e1 = self.enc1(x); e2 = self.enc2(self.pool1(e1))
        e3 = self.enc3(self.pool2(e2)); b = self.bottle(self.pool3(e3))
        d3 = self.dec3(torch.cat([self._up(self.up3, b,  e3), e3], dim=1))
        d2 = self.dec2(torch.cat([self._up(self.up2, d3, e2), e2], dim=1))
        d1 = self.dec1(torch.cat([self._up(self.up1, d2, e1), e1], dim=1))
        return self.head(d1)

# =============================================================================
# PREPROCESSING + CROP UTILS
# =============================================================================

def apply_clahe(arr):
    return cv2.createCLAHE(clipLimit=2.0, tileGridSize=(8,8)).apply(arr)

def preprocess_full(pil_gray, size):
    arr = apply_clahe(np.array(pil_gray, dtype=np.uint8))
    arr = cv2.resize(arr, (size, size))
    return torch.from_numpy(arr.astype(np.float32)/255.).unsqueeze(0).unsqueeze(0)

def preprocess_crop(crop_arr, size, mask=None):
    arr = apply_clahe(crop_arr)
    arr = cv2.resize(arr, (size, size))
    if mask is not None:
        m     = cv2.resize(mask, (size, size))
        kern  = np.ones((9,9), np.uint8)
        m_dil = cv2.dilate((m > 0.15).astype(np.uint8), kern, iterations=3).astype(np.float32)
        arr   = np.clip(arr.astype(np.float32) * m_dil, 0, 255).astype(np.uint8)
    return torch.from_numpy(arr.astype(np.float32)/255.).unsqueeze(0).unsqueeze(0)

def mask_to_bbox(mask_np, thr=0.2):
    binary = (mask_np > thr).astype(np.uint8)
    if binary.sum() == 0: return None
    _, _, stats, _ = cv2.connectedComponentsWithStats(binary)
    if len(stats) <= 1: return None
    lg = 1 + np.argmax(stats[1:, cv2.CC_STAT_AREA])
    x1 = int(stats[lg, cv2.CC_STAT_LEFT]);  y1 = int(stats[lg, cv2.CC_STAT_TOP])
    x2 = x1 + int(stats[lg, cv2.CC_STAT_WIDTH])
    y2 = y1 + int(stats[lg, cv2.CC_STAT_HEIGHT])
    return (x1, y1, x2, y2)

def fix_finger_assignment(seg_probs, thr=0.2):
    def cx(ch):
        pts = np.where(seg_probs[ch] > thr)
        return float(pts[1].mean()) if len(pts[1]) > 0 else None
    l2x, l4x = cx(0), cx(1)
    if l2x is not None and l4x is not None and l2x < l4x:
        seg_probs[[0,1]] = seg_probs[[1,0]]
    r2x, r4x = cx(2), cx(3)
    if r2x is not None and r4x is not None and r2x > r4x:
        seg_probs[[2,3]] = seg_probs[[3,2]]
    return seg_probs

def make_square_crop(arr, bbox_model, model_size, orig_size, pad_frac=0.10):
    orig_w, orig_h = orig_size
    sx = orig_w / model_size; sy = orig_h / model_size
    x1=bbox_model[0]*sx; y1=bbox_model[1]*sy
    x2=bbox_model[2]*sx; y2=bbox_model[3]*sy
    pw=(x2-x1)*pad_frac; ph=(y2-y1)*pad_frac
    x1-=pw; y1-=ph; x2+=pw; y2+=ph
    cx_=(x1+x2)/2; cy_=(y1+y2)/2; side=max(x2-x1, y2-y1)
    x1=cx_-side/2; x2=cx_+side/2; y1=cy_-side/2; y2=cy_+side/2
    cx1=int(max(0,x1)); cy1=int(max(0,y1))
    cx2=int(min(orig_w,x2)); cy2=int(min(orig_h,y2))
    if cx2<=cx1 or cy2<=cy1: return None, None
    return arr[cy1:cy2, cx1:cx2], (cx1,cy1,cx2,cy2)

# =============================================================================
# PREDICTION
# =============================================================================

def predict_tta(kp_model, crop_t, device):
    with torch.no_grad():
        h_orig = kp_model(crop_t.to(device))[0]
        h_flip = kp_model(crop_t.flip(-1).to(device))[0]
    return (h_orig + h_flip.flip(-1)) / 2

def extract_keypoints(hm_t, thr=0.0):
    hms = torch.sigmoid(hm_t).cpu().numpy()
    kps = np.full((hms.shape[0], 2), np.nan, dtype=np.float32)
    for j in range(hms.shape[0]):
        mv = hms[j].max()
        if mv < thr: continue
        y, x = np.unravel_index(hms[j].argmax(), hms[j].shape)
        kps[j] = [x / hms[j].shape[1], y / hms[j].shape[0]]
    return kps

def predict_image(orig_image, seg_model, kp_model, device):
    """Run full cascade and return dict of label → (x, y) in original image coords."""
    orig_arr  = np.array(orig_image, dtype=np.uint8)
    orig_size = orig_image.size

    seg_in = preprocess_full(orig_image, CONFIG['image_size']).to(device)
    with torch.no_grad():
        seg_p = torch.sigmoid(seg_model(seg_in))[0].cpu().numpy()
    seg_p = fix_finger_assignment(seg_p)

    masks_up = np.stack([
        cv2.resize(seg_p[c], orig_size, interpolation=cv2.INTER_LINEAR)
        for c in range(4)], axis=0)

    kp_coords = {}   # label → (x, y)

    for ch, fl in enumerate(FINGER_LABELS):
        kp_labels = FINGER_KP_NAMES[fl]
        bbox = mask_to_bbox(seg_p[ch], CONFIG['seg_threshold'])
        if bbox is None: continue

        c_arr, c_box = make_square_crop(orig_arr, bbox,
                                        CONFIG['image_size'], orig_size,
                                        CONFIG['pad_fraction'])
        if c_arr is None: continue

        cx1, cy1, cx2, cy2 = c_box
        cw = cx2 - cx1; ch_ = cy2 - cy1
        finger_mask = masks_up[ch, cy1:cy2, cx1:cx2]

        crop_t = preprocess_crop(c_arr, CONFIG['crop_size'],
                                 mask=finger_mask).to(device)
        hm     = predict_tta(kp_model, crop_t, device)
        coords = extract_keypoints(hm, CONFIG['kp_threshold'])

        for j, lbl in enumerate(kp_labels):
            if not np.isnan(coords[j, 0]):
                kp_coords[lbl] = (
                    float(coords[j, 0] * cw + cx1),
                    float(coords[j, 1] * ch_ + cy1),
                )

    return kp_coords

# =============================================================================
# EXCEL OUTPUT
# =============================================================================

FINGER_BONE_GROUPS = {
    'L2': ['L21', 'L22', 'L23'],
    'L4': ['L41', 'L42', 'L43'],
    'R2': ['R21', 'R22', 'R23'],
    'R4': ['R41', 'R42', 'R43'],
}


def build_excel(rows, output_path):
    """
    rows: list of dicts, one per image per bone, with keys:
        image, bone_id, top_label, bot_label,
        top_x, top_y, bot_x, bot_y,
        length_px, length_mm, detected
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Bone Measurements"

    # ── styles ────────────────────────────────────────────────────────────────
    hdr_fill  = PatternFill("solid", fgColor="1F4E79")
    hdr_font  = Font(bold=True, color="FFFFFF", size=10)
    grp_fills = {
        'L2': PatternFill("solid", fgColor="FFF0FF"),
        'L4': PatternFill("solid", fgColor="FFF5E6"),
        'R2': PatternFill("solid", fgColor="E6FFFF"),
        'R4': PatternFill("solid", fgColor="E6FFE6"),
    }
    total_fills = {
        'L2': PatternFill("solid", fgColor="E8C8E8"),
        'L4': PatternFill("solid", fgColor="E8D8C0"),
        'R2': PatternFill("solid", fgColor="B8E8E8"),
        'R4': PatternFill("solid", fgColor="B8E8B8"),
    }
    miss_font  = Font(color="AAAAAA", italic=True, size=9)
    num_font   = Font(size=9)
    total_font = Font(bold=True, size=9)
    thin   = Side(style='thin',   color="CCCCCC")
    medium = Side(style='medium', color="888888")
    border = Border(left=thin,   right=thin,   top=thin,   bottom=thin)
    total_border = Border(left=medium, right=medium, top=medium, bottom=medium)
    center = Alignment(horizontal='center', vertical='center')

    # ── header row ────────────────────────────────────────────────────────────
    headers = [
        'Image', 'Bone', 'Top label', 'Top X (px)', 'Top Y (px)',
        'Bot label', 'Bot X (px)', 'Bot Y (px)',
        'Length (px)', 'Length (mm)', 'Detected'
    ]
    for col, h in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = hdr_font; cell.fill = hdr_fill
        cell.alignment = center; cell.border = border

    # ── data rows + finger subtotals ──────────────────────────────────────────
    # Group rows by image then by finger
    from itertools import groupby

    r_idx = 2
    rows_by_image = {}
    for row in rows:
        rows_by_image.setdefault(row['image'], []).append(row)

    for img_name, img_rows in rows_by_image.items():
        rows_by_finger = {}
        for row in img_rows:
            rows_by_finger.setdefault(row['bone_id'][:2], []).append(row)

        for finger in ['L2', 'L4', 'R2', 'R4']:
            f_rows = rows_by_finger.get(finger, [])
            fill   = grp_fills.get(finger, PatternFill())

            for row in f_rows:
                detected = row['detected']
                values = [
                    row['image'], row['bone_id'], row['top_label'],
                    round(row['top_x'], 1)     if detected else None,
                    round(row['top_y'], 1)     if detected else None,
                    row['bot_label'],
                    round(row['bot_x'], 1)     if detected else None,
                    round(row['bot_y'], 1)     if detected else None,
                    round(row['length_px'], 2) if detected else None,
                    round(row['length_mm'], 2) if detected else None,
                    'Yes' if detected else 'No',
                ]
                for col, val in enumerate(values, start=1):
                    cell = ws.cell(row=r_idx, column=col, value=val)
                    cell.fill = fill; cell.border = border; cell.alignment = center
                    cell.font = num_font if detected else miss_font
                r_idx += 1

            # Finger subtotal row
            det_rows  = [r for r in f_rows if r['detected']]
            total_px  = sum(r['length_px'] for r in det_rows)
            total_mm  = sum(r['length_mm'] for r in det_rows)
            t_fill    = total_fills.get(finger, PatternFill())
            sub_vals  = [
                img_name, f'{finger} TOTAL', '', '', '', '', '', '',
                round(total_px, 2) if det_rows else None,
                round(total_mm, 2) if det_rows else None,
                f'{len(det_rows)}/3 bones',
            ]
            for col, val in enumerate(sub_vals, start=1):
                cell = ws.cell(row=r_idx, column=col, value=val)
                cell.fill = t_fill; cell.border = total_border
                cell.alignment = center; cell.font = total_font
            r_idx += 1

    # ── Sheet 2: Summary per bone ─────────────────────────────────────────────
    ws2 = wb.create_sheet("Summary per bone")

    bone_lengths = {b[0]: [] for b in BONES}
    for row in rows:
        if row['detected']:
            bone_lengths[row['bone_id']].append(row['length_mm'])

    hdr2 = ['Bone', 'N images', 'Mean (mm)', 'Std (mm)', 'Min (mm)', 'Max (mm)']
    for col, h in enumerate(hdr2, start=1):
        cell = ws2.cell(row=1, column=col, value=h)
        cell.font = hdr_font; cell.fill = hdr_fill
        cell.alignment = center; cell.border = border

    r2 = 2
    for finger in ['L2', 'L4', 'R2', 'R4']:
        for bone_id in FINGER_BONE_GROUPS[finger]:
            vals = bone_lengths[bone_id]
            fill = grp_fills.get(finger, PatternFill())
            row_vals = [
                bone_id, len(vals),
                round(float(np.mean(vals)), 2) if vals else None,
                round(float(np.std(vals)),  2) if vals else None,
                round(float(np.min(vals)),  2) if vals else None,
                round(float(np.max(vals)),  2) if vals else None,
            ]
            for col, val in enumerate(row_vals, start=1):
                cell = ws2.cell(row=r2, column=col, value=val)
                cell.fill = fill; cell.border = border
                cell.alignment = center; cell.font = num_font
            r2 += 1

        # Finger total summary row
        all_finger_vals = []
        for bone_id in FINGER_BONE_GROUPS[finger]:
            all_finger_vals.append(bone_lengths[bone_id])
        # Sum the 3 bones per image to get finger totals
        n_images = len(all_finger_vals[0]) if all_finger_vals[0] else 0
        finger_totals = []
        for i in range(n_images):
            s = sum(all_finger_vals[b][i] for b in range(3)
                    if i < len(all_finger_vals[b]))
            finger_totals.append(s)

        t_fill = total_fills.get(finger, PatternFill())
        t_vals = [
            f'{finger} TOTAL', len(finger_totals),
            round(float(np.mean(finger_totals)), 2) if finger_totals else None,
            round(float(np.std(finger_totals)),  2) if finger_totals else None,
            round(float(np.min(finger_totals)),  2) if finger_totals else None,
            round(float(np.max(finger_totals)),  2) if finger_totals else None,
        ]
        for col, val in enumerate(t_vals, start=1):
            cell = ws2.cell(row=r2, column=col, value=val)
            cell.fill = t_fill; cell.border = total_border
            cell.alignment = center; cell.font = total_font
        r2 += 1

    # ── Sheet 3: Finger totals per image ─────────────────────────────────────
    ws3 = wb.create_sheet("Finger Totals per image")

    hdr3 = ['Image',
            'L2 total (px)', 'L2 total (mm)',
            'L4 total (px)', 'L4 total (mm)',
            'R2 total (px)', 'R2 total (mm)',
            'R4 total (px)', 'R4 total (mm)',
            'All 4 fingers (mm)']
    for col, h in enumerate(hdr3, start=1):
        cell = ws3.cell(row=1, column=col, value=h)
        cell.font = hdr_font; cell.fill = hdr_fill
        cell.alignment = center; cell.border = border

    for r3_idx, (img_name, img_rows) in enumerate(rows_by_image.items(), start=2):
        by_bone = {r['bone_id']: r for r in img_rows}
        row_vals = [img_name]
        grand_mm = 0.0
        for finger in ['L2', 'L4', 'R2', 'R4']:
            bone_ids = FINGER_BONE_GROUPS[finger]
            det = [by_bone[b] for b in bone_ids if b in by_bone and by_bone[b]['detected']]
            total_px = sum(r['length_px'] for r in det)
            total_mm = sum(r['length_mm'] for r in det)
            grand_mm += total_mm
            row_vals += [
                round(total_px, 2) if det else None,
                round(total_mm, 2) if det else None,
            ]
        row_vals.append(round(grand_mm, 2))

        for col, val in enumerate(row_vals, start=1):
            cell = ws3.cell(row=r3_idx, column=col, value=val)
            cell.border = border; cell.alignment = center; cell.font = num_font

    # ── Sheet 4: 2D:4D Ratios ─────────────────────────────────────────────────
    ws4 = wb.create_sheet("2D4D Ratios")

    hdr4 = ['Image', 'L2 total (px)', 'L4 total (px)', 'Left 2D:4D',
             'R2 total (px)', 'R4 total (px)', 'Right 2D:4D']
    for col, h in enumerate(hdr4, start=1):
        cell = ws4.cell(row=1, column=col, value=h)
        cell.font = hdr_font; cell.fill = hdr_fill
        cell.alignment = center; cell.border = border

    ratio_data = []
    for r4_idx, (img_name, img_rows) in enumerate(rows_by_image.items(), start=2):
        by_bone = {r['bone_id']: r for r in img_rows}

        l2_px = sum(by_bone[b]['length_px'] for b in FINGER_BONE_GROUPS['L2']
                    if b in by_bone and by_bone[b]['detected'])
        l4_px = sum(by_bone[b]['length_px'] for b in FINGER_BONE_GROUPS['L4']
                    if b in by_bone and by_bone[b]['detected'])
        r2_px = sum(by_bone[b]['length_px'] for b in FINGER_BONE_GROUPS['R2']
                    if b in by_bone and by_bone[b]['detected'])
        r4_px = sum(by_bone[b]['length_px'] for b in FINGER_BONE_GROUPS['R4']
                    if b in by_bone and by_bone[b]['detected'])

        left_ratio  = round(l2_px / l4_px, 4) if l4_px > 0 else None
        right_ratio = round(r2_px / r4_px, 4) if r4_px > 0 else None

        ratio_data.append((left_ratio, right_ratio))

        row_vals = [img_name,
                    round(l2_px, 2), round(l4_px, 2), left_ratio,
                    round(r2_px, 2), round(r4_px, 2), right_ratio]
        for col, val in enumerate(row_vals, start=1):
            cell = ws4.cell(row=r4_idx, column=col, value=val)
            cell.border = border; cell.alignment = center; cell.font = num_font

    # Summary stats rows
    n_rows = len(rows_by_image) + 2
    for label, func in [('Mean', np.mean), ('Std', np.std),
                         ('Min', np.min), ('Max', np.max)]:
        left_vals  = [d[0] for d in ratio_data if d[0] is not None]
        right_vals = [d[1] for d in ratio_data if d[1] is not None]
        stat_row = [label, '', '', round(float(func(left_vals)), 4) if left_vals else None,
                    '', '', round(float(func(right_vals)), 4) if right_vals else None]
        for col, val in enumerate(stat_row, start=1):
            cell = ws4.cell(row=n_rows, column=col, value=val)
            cell.fill = hdr_fill; cell.border = total_border
            cell.alignment = center; cell.font = total_font
        n_rows += 1

    # ── column widths ─────────────────────────────────────────────────────────
    for ws_ in [ws, ws2, ws3, ws4]:
        for col in ws_.columns:
            max_len = max((len(str(c.value)) if c.value is not None else 0) for c in col)
            ws_.column_dimensions[get_column_letter(col[0].column)].width = max(max_len + 3, 12)

    wb.save(output_path)

# =============================================================================
# MAIN
# =============================================================================

def main():
    device = torch.device('cuda' if torch.cuda.is_available() else 'cpu')

    for p in [CONFIG['seg_model_path'], CONFIG['kp_model_path']]:
        if not os.path.exists(p):
            print(f"ERROR: {p} not found."); return

    seg_ck = torch.load(CONFIG['seg_model_path'], map_location=device)
    seg_model = FingerSegUNet(4).to(device)
    seg_model.load_state_dict(seg_ck['model_state_dict']); seg_model.eval()

    kp_ck = torch.load(CONFIG['kp_model_path'], map_location=device)
    kp_model = FingerKpModel(6).to(device)
    kp_model.load_state_dict(kp_ck['model_state_dict']); kp_model.eval()
    print("Models loaded.\n")

    image_files = sorted(Path(CONFIG['input_folder']).glob('*.png')) + \
                  sorted(Path(CONFIG['input_folder']).glob('*.PNG')) + \
                  sorted(Path(CONFIG['input_folder']).glob('*.jpg')) + \
                  sorted(Path(CONFIG['input_folder']).glob('*.jpeg'))
    image_files = sorted(set(image_files))

    if not image_files:
        print(f"No images found in {CONFIG['input_folder']}"); return

    print(f"Found {len(image_files)} images. Running predictions...\n")

    all_rows = []

    for img_path in image_files:
        print(f"  {img_path.name} ...", end=' ')
        try:
            orig_image = Image.open(img_path).convert('L')
            kp_coords  = predict_image(orig_image, seg_model, kp_model, device)

            for bone_id, top_lbl, bot_lbl in BONES:
                top = kp_coords.get(top_lbl)
                bot = kp_coords.get(bot_lbl)
                detected = top is not None and bot is not None

                if detected:
                    length_px = float(np.hypot(bot[0]-top[0], bot[1]-top[1]))
                    length_mm = length_px * CONFIG['px_to_mm']
                else:
                    length_px = length_mm = 0.0

                all_rows.append({
                    'image':     img_path.name,
                    'bone_id':   bone_id,
                    'top_label': top_lbl,
                    'bot_label': bot_lbl,
                    'top_x':     top[0] if top else 0.0,
                    'top_y':     top[1] if top else 0.0,
                    'bot_x':     bot[0] if bot else 0.0,
                    'bot_y':     bot[1] if bot else 0.0,
                    'length_px': length_px,
                    'length_mm': length_mm,
                    'detected':  detected,
                })

            detected_bones = sum(1 for b in all_rows[-len(BONES):]
                                 if b['detected'])
            mean_len = np.mean([b['length_mm'] for b in all_rows[-len(BONES):]
                                if b['detected']]) if detected_bones else 0
            print(f"{detected_bones}/12 bones  |  mean length {mean_len:.1f} mm")

        except Exception as e:
            import traceback
            print(f"ERROR: {e}"); traceback.print_exc()
            for bone_id, top_lbl, bot_lbl in BONES:
                all_rows.append({
                    'image': img_path.name, 'bone_id': bone_id,
                    'top_label': top_lbl, 'bot_label': bot_lbl,
                    'top_x': 0., 'top_y': 0., 'bot_x': 0., 'bot_y': 0.,
                    'length_px': 0., 'length_mm': 0., 'detected': False,
                })

    build_excel(all_rows, CONFIG['output_file'])

    total   = len(image_files)
    det_all = sum(1 for r in all_rows if r['detected'])
    print(f"\nDone. {total} images  |  {det_all}/{total*len(BONES)} bone measurements detected.")
    print(f"Saved: {CONFIG['output_file']}")


if __name__ == '__main__':
    try:
        main()
    except Exception:
        import traceback; traceback.print_exc()
